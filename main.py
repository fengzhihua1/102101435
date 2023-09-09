import requests
import re
import heapq
import openpyxl
counts = {}
header = {
    'cookie':'buvid3=F6C66E54-FB1F-179C-B44E-7F5757CDB2BD44709infoc; b_nut=1693794644; i-wanna-go-back=-1; b_ut=7; _uuid=194C10C2A-10D910-EB3E-A223-A49A2BB7910FF44443infoc; buvid_fp=da33d3972c60dc5743b8faa4fc8bf7c7; CURRENT_FNVAL=4048; rpdid=|(u))kkYu|uJ0J\'uYmJ~mlRJ~; header_theme_version=CLOSE; buvid4=52303CC1-28D6-6DFA-305C-FB58923AA91D45519-023090410-%2FxwqHe8zHTV4wPiaY2mhbA%3D%3D; bili_ticket=eyJhbGciOiJIUzI1NiIsImtpZCI6InMwMyIsInR5cCI6IkpXVCJ9.eyJleHAiOjE2OTQyMjExMTksImlhdCI6MTY5Mzk2MTkxOSwicGx0IjotMX0.so8ypMIKhWq0qpqehDVf3aFb7BYffuJGWlnuFJRUexM; bili_ticket_expires=1694221119; SESSDATA=5c756941%2C1709518904%2Cfa7fd%2A91JJUf1UoDYTc8IRywHIz49aPwwY6NAnu9PXuJNtqQuEvGyEH0et-RDOOQned_wJaGxtY4yQAABAA; bili_jct=9fd323eedad65c235fe4db3314fb6f34; DedeUserID=324839604; DedeUserID__ckMd5=a4494cfea0536aca; sid=6enlped9; PVID=1; b_lsid=FBABB623_18A686CE976; bsource=search_baidu; innersign=0; home_feed_column=5; browser_resolution=1495-708; bp_video_offset_324839604=837867271252606993',
    'user-agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/101.0.4951.64 Safari/537.36'
}
def get_html(i):#爬取搜索页网址

    url = "https://search.bilibili.com/all?keyword=日本核污染排海"
    if i != 0:
        url += '&page=' + str(i+1)
    content = requests.get(url, headers=header).text
    return content

def get_bv_number(content):#获取视频的BV号
    bv = re.findall('bvid:"(.*?)",title', content)
    return bv

def get_msg_adr(bv,idx,cnt):#获取每个视频的弹幕地址
    print('正在爬取第%d个视频'%(cnt+idx+1));
    url = 'https://www.ibilibili.com/video/' + bv[idx]
    response = requests.get(url).text
    address = re.findall('<a href="(.*?)"  class="btn btn-default" target="_blank">弹幕</a>', response)
    return address

def get_msg(adress):#获取每个视频的弹幕
    response = requests.get(url=adress[0])
    response.encoding = 'utf-8'
    html_data = re.findall('<d p=".*?">(.*?)</d>', response.text)
    return html_data

def to_txt_record(html_data):#记录弹幕
    for msg in html_data:
        with open('msg.txt', mode='a', encoding='utf-8') as f:
            f.write(msg)
            f.write('\n')
            if msg in counts:
                counts[msg] += 1
            else:
                counts[msg] = 1

def top20_message(data_dict,N):#获取排名前20的弹幕
    dic={}
    dic.update({k: data_dict[k] for k in heapq.nlargest(N, data_dict, key=data_dict.get)})
    arr = []
    for key,value in dic.items():
        arr.append((key,value))
    return arr

def write_lines_excel(arr):#将排名前20的弹幕写入表格
    work_book = openpyxl.Workbook()
    sheet = work_book.create_sheet('top20弹幕概况')
    sheet.cell(1, 1, '数量排名前20的弹幕')
    sheet.cell(1, 2, '个数')
    for index, row in enumerate(arr):
        for col in range(len(row)):
            sheet.cell(index + 2, col + 1, row[col])
    work_book.save('Top20.xlsx')
print('请输入爬取视频的数量：')
video_num=input();#输入需要爬取视频的数量;
video_num=int(video_num)
cnt = 0
for i in range (0, 1000):
    if video_num<=0:
        break;
    html = get_html(i)
    BV = get_bv_number(html)
    for idx in range(0,42):
        adress = get_msg_adr(BV, idx,cnt)
        html_data = get_msg(adress)
        to_txt_record(html_data)
        video_num-=1
        if video_num==0:break
    cnt+=42
print('爬取完毕')
arr = top20_message(counts,20)
write_lines_excel(arr)
