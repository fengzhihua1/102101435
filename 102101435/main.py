import requests
import re
import heapq
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import urllib
import time
counts = {}
def get_bv_numbers(keyword: str, bv_number_count: int):
    print(f'开始爬取有关 {keyword} 的视频弹幕')
    options = Options()
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36 Edg/116.0.1938.69")
    options.add_argument('--headless')
    driver = webdriver.Edge(options=options)
    driver.maximize_window()
    edge_options = webdriver.EdgeOptions()
    edge_options.add_argument("--disable-application-cache")
    edge_options.add_argument("--disable-gpu-shader-disk-cache")
    edge_options.add_argument("--disable-gpu-program-cache")
    edge_options.add_argument("--disable-cache")

    encoded_keyword = urllib.parse.quote(keyword)
    multiplier = 0
    res = set()
    while len(res) < bv_number_count:
        url = "https://search.bilibili.com/all?keyword=" + encoded_keyword + f'&o={24 * multiplier}'
        driver.get(url)
        driver.refresh()
        time.sleep(10)  # 保证页面加载完全
        element_content = driver.page_source
        element_content = element_content[element_content.find(r'video-list row'): element_content.find(r'to_hide_xs')]
        pattern = r'BV(.*?)/"'
        bvids = re.findall(pattern, element_content)
        res.update(bvids[: bv_number_count - len(res)])
        multiplier += 1
    driver.quit()
    print("获取完全")
    print(res)
    return list(res)


def get_msg_adr(bv):#获取每个视频的弹幕地址
    url = 'https://www.ibilibili.com/video/BV' + bv
    response = requests.get(url).text
    address = re.findall('<a href="(.*?)"  class="btn btn-default" target="_blank">弹幕</a>', response)
    return address

def get_msg(adress):#获取每个视频的弹幕
    response = requests.get(url=adress[0])
    response.encoding = 'utf-8'
    html_data = re.findall('<d p=".*?">(.*?)</d>', response.text)
    return html_data

def to_txt_record(html_data):#记录弹幕
    for msg in html_data:
        with open('msg.txt', mode='a', encoding='utf-8') as f:
            f.write(msg)
            f.write('\n')
            if msg in counts:
                counts[msg] += 1
            else:
                counts[msg] = 1

def top20_message(data_dict,N):#获取排名前20的弹幕
    dic={}
    dic.update({k: data_dict[k] for k in heapq.nlargest(N, data_dict, key=data_dict.get)})
    arr = []
    for key,value in dic.items():
        arr.append((key,value))
    return arr

def write_lines_excel(arr):#将排名前20的弹幕写入表格
    work_book = openpyxl.Workbook()
    sheet = work_book.create_sheet('top20弹幕概况')
    sheet.cell(1, 1, '数量排名前20的弹幕')
    sheet.cell(1, 2, '个数')
    for index, row in enumerate(arr):
        for col in range(len(row)):
            sheet.cell(index + 2, col + 1, row[col])
    work_book.save('Top20.xlsx')
print("要搜索的关键词：")
str = input()
print('请输入爬取视频的数量：')
video_num=input();#输入需要爬取视频的数量;
video_num=int(video_num)
BV = get_bv_numbers(str,video_num)
for bv in BV:
    addr=get_msg_adr(bv)
    html_data=get_msg(addr)
    to_txt_record(html_data)
print('爬取完毕')
arr = top20_message(counts,20)
write_lines_excel(arr)
